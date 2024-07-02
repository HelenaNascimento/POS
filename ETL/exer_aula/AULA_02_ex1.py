import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet,  ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors

#carregando arquivo do excel

arquivo_excel = "/home/technical/WorkSpace/POS/cursos.xlsx"
wb  = openpyxl.load_workbook(arquivo_excel)
ws = wb.active

#Obtendo dados da planilha
#linha de cabeçalho

titulos_colunas = [celula.value for celula in ws[1]]

dados = []

for linha in ws.iter_rows(min_row=2):
    #linha de dados
    dados_linha = [celula.value for celula in linha]
    dados.append(dados_linha)
    
#criando o PDF

nome_PDF =  "Cursos_Univ.pdf"
pdf = SimpleDocTemplate(nome_PDF, pagesize=landscape(letter))
estilo = getSampleStyleSheet()

#Cabeçalho

estilo_cabecalho = ParagraphStyle("Cabecalho", parent = estilo["Heading1"], textColor=colors.red  )
cabecalho = Paragraph("Tabela", estilo_cabecalho)

#Tabela

tabela =  [titulos_colunas] + dados

#Adcionar o cabeçalho e a tabela ao conteudo 

estilo_tabela = TableStyle([('BACKGROUP', (2,0), (2,0), '#CCCCCC')])


tabela_pdf = Table(tabela)
tabela_pdf.setStyle(estilo_tabela)

conteudo = [cabecalho, tabela_pdf]

#print(conteudo)

pdf.build(conteudo)
